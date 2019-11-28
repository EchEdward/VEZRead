from openpyxl import load_workbook, Workbook


def OpenFile(path):
    Katal = load_workbook(filename = path)
    sheets = Katal.sheetnames
    names = []
    N = []
    t = []
    lv =[]
    for sheet in sheets:
        Kat = Katal[sheet]
        i=1
        while Kat["A"+str(i)].value != None:
            if str(type(Kat["A"+str(i)].value)) == "<class 'str'>" or (Kat["A"+str(i)].value!=None and Kat["B"+str(i)].value==None and Kat["C"+str(i)].value==None):
                n = str(Kat["A"+str(i)].value)
                vli = Kat["B"+str(i)].value
                ti = Kat["C"+str(i)].value
                i+=1
                a = []
                while str(type(Kat["A"+str(i)].value)) != "<class 'str'>" and Kat["A"+str(i)].value != None:
                    a1 = Kat["A"+str(i)].value if Kat["A"+str(i)].value != None else ''
                    a2 = Kat["B"+str(i)].value if Kat["B"+str(i)].value != None else ''
                    a3 = Kat["C"+str(i)].value if Kat["C"+str(i)].value != None else ''
                    a.append([a1,a2,a3])
                    if (Kat["A"+str(i)].value!=None and Kat["B"+str(i)].value==None and Kat["C"+str(i)].value==None):
                        i+=1
                        break
                    i+=1
                i-=1
                if a != []:
                    names.append(n)
                    N.append(a)
                    lv.append(vli)
                    t.append(ti)
                    
            i+=1

    return N, names, lv, t

def SaveFile1(fname,nam,N,lv,t):
    wb = Workbook()
    ws=[]
    ws.append(wb.active)
    #ws[0].append(['№','Км','A','B','C','Трос'])
    for i in range(len(nam)):
        ws[0].append([nam[i],lv[i],t[i]])
        for j in range(len(N[i])):
            ws[0].append([N[i][j][0],N[i][j][1] if N[i][j][1]!='' else None, N[i][j][2] if N[i][j][2]!='' else None])
    wb.save(fname)


def SaveFile2(fname,nam,pe1,pe2,pe,lv,t):
    wb = Workbook()
    ws=[]
    ws.append(wb.active)
    ws[0].append(['Название','Pэ1, Ом','Pэ2, Ом','Pэ, Ом','lв, м','t, м'])
    for i in range(len(nam)):
        ws[0].append([nam[i],pe1[i],pe2[i],pe[i],lv[i],t[i]])
        
    wb.save(fname)



