import numpy as np
from scipy import interpolate
from openpyxl import load_workbook
import copy

Katal = load_workbook(filename = 'Lines/3.xlsx')
#print(Katal.sheetnames)
Kat = Katal['3']

i=1
X3 = []
Y3 = []
while Kat["A"+str(i)].value != None:
    X3.append(Kat["A"+str(i)].value)
    Y3.append(Kat["B"+str(i)].value)
    i+=1


f3 = interpolate.interp1d(X3, Y3)


#arr = [[7030.0, 1.4, 1.4], [5957.0, 1.26, 2.66], [926.0, 3.18, 5.84], [271.0, '', '']]
#arr = [[7499.0, 0.874, 0.874], [7080.0, 1.08, 1.96], [5624.0, 1.27, 3.23], [1289.0, 4.27, 7.5], [198.0, '', '']]
#arr = [[5746.0, 1.01, 1.01], [6889.0, 1.2, 2.21], [6262.0, 0.365, 2.58], [2107.0, 0.572, 3.15], [935.0, '', '']]

def RschRoo(arr1,lv,t):
    arr = copy.deepcopy(arr1)
    hr = lv*1.4
    he1 = lv+t
    h = 0
    ind =-1
    i=-1
    trig = False
    while i<len(arr):
        i+=1
        h+=arr[i][1] if arr[i][1]!='' else 1000
        if h==he1 and not trig:
            ind = i
            trig =True
        elif h>he1 and not trig:
            a = [arr[i][0],h-he1,h]
            b = [arr[i][0],(arr[i][1] if arr[i][1]!='' else 1000)-a[1],he1]
            arr = arr[0:i]+[b]+[a]+arr[i+1:]
            ind = i
            i+=1
            trig =True
        if h>=hr:
            a = [arr[i][0],(arr[i][1] if arr[i][1]!='' else 1000)-(h-hr),hr]
            arr = arr[:i]+[a]
            break

    s = 0
    for i in range(ind+1):
        s+=arr[i][1]/arr[i][0]
    p1 = he1/s

    if len(arr)-(ind+1)==1:
        p2 = arr[ind+1][0]
    else:
        s = 0
        trig = arr[ind+1][0]>arr[len(arr)-1][0]
        for i in range(ind+1,len(arr)):
            s += arr[i][1]*arr[i][0] if trig else arr[i][1]/arr[i][0]
        p2 = s/(hr-he1) if trig else (hr-he1)/s

    p = p2*f3(p1/p2)
    
    return p1,p2,p

            
#print(RschRoo(arr,5,0.5)) 
    
    